import io, os, csv, base64, re
import pandas as pd
from flask import Flask, request, Response
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)

HTML_PAGE = """
<!doctype html>
<html lang="pt">
  <head>
    <meta charset="utf-8" />
    <title>Excel → Odoo CSV</title>
    <style>
      body { font-family: system-ui, sans-serif; margin: 3rem; background:#fafafa; }
      .card { max-width: 700px; margin: auto; background:white; padding:2rem; border-radius:12px; box-shadow:0 4px 15px rgba(0,0,0,.08); }
      h1 { margin-top: 0; }
      input[type=file] { width:100%; padding:.7rem; border:1px solid #ccc; border-radius:6px; margin-top:.5rem; }
      button { margin-top:1rem; padding:.8rem 1.5rem; background:#111827; color:white; border:0; border-radius:8px; cursor:pointer; font-weight:600; }
    </style>
  </head>
  <body>
    <div class="card">
      <h1>Excel → CSV (Odoo Base64)</h1>
      <form action="/convert" method="post" enctype="multipart/form-data">
        <label for="file">Seleciona um ficheiro Excel (.xlsx)</label>
        <input id="file" name="file" type="file" accept=".xlsx" required />
        <button type="submit">Converter</button>
      </form>
    </div>
  </body>
</html>
"""

def excel_to_odoo_csv(xlsx_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    # Extrair imagens e converter em Base64
    image_map = {}
    for image in getattr(ws, "_images", []):
        try:
            if hasattr(image.anchor, "_from"):
                col = image.anchor._from.col + 1
                row = image.anchor._from.row + 1
                coord = ws.cell(row=row, column=col).coordinate
                img_bytes = image._data()
                image_map[coord] = base64.b64encode(img_bytes).decode("utf-8")
        except Exception:
            continue

    # Ler dados (valores, não fórmulas)
    data = list(ws.values)
    if not data or not data[0]:
        raise ValueError("O ficheiro não contém dados válidos.")
    columns = data[0]
    rows = data[1:]
    df = pd.DataFrame(rows, columns=columns)

    # Substituir coluna IMAGE pelos Base64
    if "IMAGE" in df.columns:
        col_idx = list(df.columns).index("IMAGE") + 1
        col_letter = get_column_letter(col_idx)
        for i in range(len(df)):
            cell_ref = f"{col_letter}{i + 2}"
            if cell_ref in image_map:
                df.at[i, "IMAGE"] = image_map[cell_ref]

    # Exportar CSV compatível com Odoo
    buf = io.StringIO()
    df.to_csv(buf, index=False, header=True, sep=";", quoting=csv.QUOTE_ALL, encoding="utf-8")
    csv_data = buf.getvalue()

    # --- LIMPEZA FINAL ---
    cleaned_lines = []
    current_line = ""
    for raw_line in csv_data.splitlines():
        line = raw_line.rstrip("\n").rstrip("\r")
        if line.startswith('"') or line.startswith("Ref n."):
            if current_line:
                cleaned_lines.append(current_line)
            current_line = line
        else:
            current_line += " " + line.strip()
    if current_line:
        cleaned_lines.append(current_line)

    # Limpar espaços dentro de campos entre aspas e remover #VALUE!
    final_lines = []
    for line in cleaned_lines:
        # remove "#VALUE!"
        line = line.replace("#VALUE!", "")
        # remove espaços após aspas de abertura
        line = re.sub(r'";\s*"', '";"', line)
        line = re.sub(r'"\s+', '"', line)
        final_lines.append(line)

    cleaned_csv = "\n".join(final_lines)
    return cleaned_csv.encode("utf-8")

@app.get("/")
def index():
    return Response(HTML_PAGE, mimetype="text/html; charset=utf-8")

@app.post("/convert")
def convert():
    if "file" not in request.files:
        return ("Nenhum ficheiro enviado.", 400)
    f = request.files["file"]
    try:
        csv_bytes = excel_to_odoo_csv(f.read())
    except Exception as e:
        return (f"Erro ao converter: {e}", 400)
    out_name = os.path.splitext(f.filename or "export.xlsx")[0] + "_odoo.csv"
    return Response(
        csv_bytes,
        headers={
            "Content-Type": "text/csv; charset=utf-8",
            "Content-Disposition": f"attachment; filename={out_name}",
        },
        status=200,
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "8000"))
    app.run(host="0.0.0.0", port=port, debug=False)
